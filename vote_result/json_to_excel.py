import json
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# === 設定區 ===
INPUT_FOLDER = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILENAME = "16族發音對照總表.xlsx"

# 16族名稱對照表 (根據編號自動填入表頭)
TRIBE_NAMES = {
    "01": "阿美語", "02": "泰雅語", "03": "排灣語", "04": "布農語",
    "05": "卑南語", "06": "魯凱語", "07": "鄒語",   "08": "賽夏語",
    "09": "雅美語", "10": "邵語",   "11": "噶瑪蘭語", "12": "太魯閣語",
    "13": "撒奇萊雅語", "14": "賽德克語", "15": "拉阿魯哇語", "16": "卡那卡那富語"
}

# 拼音 -> 注音 對照表
PINYIN_TO_ZHUYIN = {
    # 聲母
    "b": "ㄅ", "p": "ㄆ", "m": "ㄇ", "f": "ㄈ",
    "d": "ㄉ", "t": "ㄊ", "n": "ㄋ", "l": "ㄌ",
    "g": "ㄍ", "k": "ㄎ", "h": "ㄏ",
    "j": "ㄐ", "q": "ㄑ", "x": "ㄒ",
    "zh": "ㄓ", "ch": "ㄔ", "sh": "ㄕ", "r": "ㄖ",
    "z": "ㄗ", "c": "ㄘ", "s": "ㄙ", "0c": "零聲母",
    # 韻母
    "a": "ㄚ", "o": "ㄛ", "ㄜ": "ㄜ","e": "ㄝ",
    "ai": "ㄞ", "ei": "ㄟ", "ao": "ㄠ", "ou": "ㄡ",
    "an": "ㄢ", "en": "ㄣ", "ang": "ㄤ", "eng": "ㄥ", "er": "ㄦ",
    "i": "ㄧ", "ia": "ㄧㄚ", "io": "ㄧㄛ", "ie": "ㄧㄝ", "iai": "ㄧㄞ",
    "iao": "ㄧㄠ", "iou": "ㄧㄡ", "ian": "ㄧㄢ", "in": "ㄧㄣ",
    "iang": "ㄧㄤ", "ing": "ㄧㄥ",
    "u": "ㄨ", "ua": "ㄨㄚ", "uo": "ㄨㄛ", "uai": "ㄨㄞ", "uei": "ㄨㄟ",
    "uan": "ㄨㄢ", "un": "ㄨㄣ", "uang": "ㄨㄤ", "ong": "ㄨㄥ","eng": "ㄨㄥ",
    "iong": "ㄩㄥ","ㄩ": "ㄩ", "ㄩn": "ㄩㄣ", "ㄩan": "ㄩㄢ", "ㄩe": "ㄩㄝ", "0v": "零韻母"
}

# 排序順序 (依照您的 vote.py)
ORDER_LIST = [
    "b", "p", "m", "f", "d", "t", "n", "l",
    "g", "k", "h", "j", "q", "x", "zh", "ch",
    "sh", "r", "z", "c", "s", "0c",
    "an", "en", "ang", "eng", "er",
    "i", "u", "ㄩ",
    "a", "o", "ㄜ", "e",
    "ai", "ei", "ao", "ou",
    "ia", "io", "ie", "iai", "iao", "iou",
    "ian", "in", "iang", "ing",
    "ua", "uo", "uai", "uei", "uan", "un", "uang", "ong",
    "ㄩe", "ㄩan", "ㄩn", "iong", "0v"
]

def main():
    print("🚀 程式開始執行...")
    
    # 1. 搜尋檔案
    files = glob.glob(os.path.join(INPUT_FOLDER, "*_output_alignment_voted.json"))
    files.sort()
    
    if not files:
        print("❌ 找不到任何 JSON 檔案！請確認檔案位置。")
        return

    print(f"📂 找到 {len(files)} 個檔案 (注意：若少於16個代表有缺漏)")

    # 2. 準備資料結構
    data_matrix = {k: {} for k in ORDER_LIST}
    tribe_ids = []

    # --- 讀取檔案迴圈 ---
    for filepath in files:
        filename = os.path.basename(filepath)
        print(f"   -> 正在讀取: {filename} ...", end="") # 顯示進度
        
        try:
            tid = filename.split('_')[0]
            tribe_ids.append(tid)

            with open(filepath, 'r', encoding='utf-8') as f:
                local_json = json.load(f)

            for pinyin_key in ORDER_LIST:
                if pinyin_key not in data_matrix:
                    data_matrix[pinyin_key] = {}

                if pinyin_key in local_json:
                    targets = local_json[pinyin_key]
                    if targets:
                        best_char = list(targets.keys())[0]
                        note = targets[best_char].get("note", "")
                        data_matrix[pinyin_key][tid] = {"char": best_char, "note": note}
                    else:
                        data_matrix[pinyin_key][tid] = {"char": "-", "note": ""}
                else:
                    data_matrix[pinyin_key][tid] = {"char": "", "note": ""}
            print(" OK") # 讀取成功
            
        except Exception as e:
            print(f" ❌ 錯誤: {e}")

    print("📊 資料讀取完畢，正在建立 Excel 表格...")

    # 3. 建立 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "16族發音總表"

    # --- 樣式 ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
    zhuyin_font = Font(bold=True, color="C00000")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    highlight_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

    # --- 寫入表頭 ---
    headers = ["注音", "拼音"]
    for tid in tribe_ids:
        name = TRIBE_NAMES.get(tid, f"族語{tid}")
        headers.append(f"{tid}\\n{name}")

    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # --- 寫入內容 ---
    row_idx = 2
    for pinyin in ORDER_LIST:
        zhuyin = PINYIN_TO_ZHUYIN.get(pinyin, "")
        row_data = [zhuyin, pinyin]
        
        for tid in tribe_ids:
            cell_data = data_matrix[pinyin].get(tid, {"char": "", "note": ""})
            char = cell_data["char"]
            note = cell_data["note"]
            row_data.append(char)
        
        ws.append(row_data)
        
        # 樣式設定
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center_align
            cell.border = thin_border

            # 如果是注音，設置為紅色
            if col_idx == 1:
                cell.font = zhuyin_font

            # 如果 note 為指定文字，塗上淺色
            if col_idx > 2 and data_matrix[pinyin][tribe_ids[col_idx - 3]]["note"] == "Auto-filled from Global (Checked Dictionary)":
                cell.fill = highlight_fill

        row_idx += 1

    # --- 調整欄寬 ---
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    for i in range(3, len(headers) + 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = 12

    ws.freeze_panes = "C2"

    # 4. 儲存檔案
    output_path = os.path.join(INPUT_FOLDER, OUTPUT_FILENAME)
    print(f"💾 正在儲存檔案至: {OUTPUT_FILENAME} ...")
    
    try:
        wb.save(output_path)
        print(f"\n✅✅✅ 成功！檔案已產生: {OUTPUT_FILENAME}")
    except PermissionError:
        print(f"\n❌❌❌ 失敗：無法寫入檔案！")
        print(f"原因：'{OUTPUT_FILENAME}' 正被 Excel 開啟中。")
        print("解決方法：請關閉 Excel 視窗後，再重新執行一次程式。")
    except Exception as e:
        print(f"\n❌ 發生未預期的錯誤: {e}")

if __name__ == "__main__":
    main()