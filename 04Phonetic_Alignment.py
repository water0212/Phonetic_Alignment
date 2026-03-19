import numpy as np
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --- 路徑設定 ---
file_path = os.path.dirname(os.path.abspath(__file__))

# 輸入資料夾 (來自上一步驟的輸出)
INPUT_FOLDER = os.path.join(file_path,'Ailgn_syllables', 'Aligned_Results')
# 輸出資料夾
OUTPUT_FOLDER = os.path.join(file_path, 'Refined_Excel')

class PhoneticAligner:
    def __init__(self):
        # 定義聲母群組 (用於模糊比對聲母相似度)
        self.consonant_chinese_groups = [
            {'b', 'p'},
            {'m'},
            {'f'}, 
            {'d', 't'}, 
            {'n'}, 
            {'l', 'r'},
            {'g', 'k'}, 
            {'d', 'j', 'z'}, 
            {'h'}, {'j'}, 
            {'x'},
            {'t','q','c'}, 
            {'z', 'c', 's', 'zh', 'ch', 'sh','x'},#新加入chinese的x，因為在對齊過程中發現這x在某些情況下與族語的z/c/s有相似性} 
            {'d', 'j', 'z', 't', 'q', 'c'}, 
            {'f'}, 

        ]
        self.consonant_tsou_groups = [
            {'b', 'p'}, 
            {'m'}, 
            {'f'}, 
            {'d', 't'}, 
            {'n'}, 
            {'l', 'r'},
            {'g', 'k', 'q', '’', '^'}, 
            {'d', 'j', 'z'}, {'h'}, 
            {'j'}, 
            {'x'},
            {'t','c'}, 
            {'z', 'c', 's','x'}, #新加入tsou的s和x，因為在對齊過程中發現這兩個字母在某些情況下與中文的z/c/s/zh/ch/sh有相似性} 
            {'d', 'j', 'z', 't', 'c'},
            {'f','v', 'b'}
        ]
        self.vowel_map = {'w': 'u', 'y': 'i'}
        self.vowel__chinese_groups = [
            {'a'},
            {'e'},
            {'ㄜ'}, 
            {'i'},
            {'ㄩ'},
            {'o'}, #新加入o，因為在對齊過程中發現這個韻母在某些情況下與族語的u有相似性} 
            {'y','e'},
        ]
        self.vowel__tsou_groups = [
            {'a'},
            {'é'},
            {'o', 'e','u','é'}, 
            {'ɨ','ʉ', 'y'},
            {'i','u'},
            {'u'}, #新加入u，因為在對齊過程中發現這個韻母在某些情況下與中文的o有相似性}
            {'y','e'}
        ]
    def reconstruct_tsou_word(self, alignment_result):
            """
            將對齊結果中的族語音節合併回原始單字
            過濾規則：
            1. 忽略 None (中文缺失的情況通常族語還在，但若族語也是 None 則忽略)
            2. 移除 '0c' (無聲母)
            3. 移除 '0v' (無韻母)
            """
            reconstructed_word = ""
            
            for item in alignment_result:
                # item 結構: (ch_syllable, ts_syllable, status)
                ts_syl = item[1]
                
                if ts_syl:
                    # 取得聲母與韻母，預設為空字串以免報錯
                    initial = ts_syl.get('initial', '')
                    final = ts_syl.get('final', '')
                    
                    # 過濾 0c, 0v
                    if initial == '0c': initial = ''
                    if final == '0v': final = ''
                    
                    # 組合
                    reconstructed_word += initial + final
                    
            return reconstructed_word

    def calculate_consonant_score(self, c_ch, c_ts):
        if c_ch == "0c" and c_ts == "0c": return 1
        if c_ch == "0c" or  c_ts == "0c": return 0

        max_score = 0
        for group_ch, group_ts in zip(self.consonant_chinese_groups, self.consonant_tsou_groups):
            if c_ch in group_ch and c_ts in group_ts:
                if c_ch == c_ts:
                    current_score = 1
                elif group_ch == {'d', 'j', 'z', 't', 'q', 'c'}:
                    current_score = 0.7
                elif group_ch == {'z', 'c', 's', 'zh', 'ch', 'sh'}:
                    current_score = 10
                else:
                    current_score = 0.8
                
                current_score = max(current_score, 0)
                if current_score > max_score:
                    max_score = current_score
        return max_score

    def dice_coefficient(self, s1, s2):
        
        def normalize(s): return "".join(self.vowel_map.get(c, c) for c in s.lower())
        n1, n2 = normalize(s1), normalize(s2)
        len1, len2 = len(n1), len(n2)
        
        if len1 == 0 or len2 == 0: return 0.0

        dp = [[0] * (len2 + 1) for _ in range(len1 + 1)]
        for i in range(1, len1 + 1):
            for j in range(1, len2 + 1):
                    dp[i][j] = max(dp[i - 1][j], dp[i][j - 1], dp[i - 1][j - 1] + self.calculate_vowel_score(n1[i - 1], n2[j - 1]))
        intersection = dp[len1][len2]
        return 2.0 * intersection / (len1 + len2)
    
    def calculate_vowel_score(self, v_ch, v_ts):
        if v_ch == "0v" and v_ts == "0v": return 1
        if v_ch == "0v" or v_ts == "0v": return 0
        if v_ch == v_ts:
            return 1.0
        max_score = 0
        for group_ch, group_ts in zip(self.vowel__chinese_groups, self.vowel__tsou_groups):
            
            if v_ch in group_ch and v_ts in group_ts:
                if v_ch == 'ㄩ':
                    current_score = 0.5
                elif group_ch == {"o"}:
                    current_score = 2
                else:
                    current_score = 0.8
                current_score = max(current_score, 0)
                if current_score > max_score:
                    max_score = current_score
        return max_score
    
    def calculate_similarity(self, syl_ch, syl_in):
        
        score_initial = self.calculate_consonant_score(syl_ch['initial'], syl_in['initial']) * 1
        score_final = self.dice_coefficient(syl_ch['final'], syl_in['final']) * 1
        #例外處理，若是0c對0c，但韻母完全不同，則視為完全不匹配
        if syl_ch['initial'] == "0c" and syl_in['initial'] == "0c" and score_final < 0.3:
            return 0
        return score_initial + score_final

    def align(self, ch_syllables, in_syllables):
        n = len(ch_syllables)
        m = len(in_syllables)
        gap_penalty = 0 
        
        # 初始化為負無窮大或極小值，防止未經許可的路徑被選擇
        dp = np.full((n + 1, m + 1), -9999.0)
        dir_matrix = [["" for _ in range(m + 1)] for _ in range(n + 1)]
        
        # 起點
        dp[0][0] = 0

        # 初始化第一列 (垂直移動 ↑)：根據要求禁止垂直，除起點外設為極小值
        for i in range(1, n + 1):
            dp[i][0] = -9999.0 
            dir_matrix[i][0] = "↑"

        # 初始化第一行 (水平移動 ←)：允許水平
        for j in range(1, m + 1):
            dp[0][j] = dp[0][j-1] + gap_penalty
            dir_matrix[0][j] = "←"
            
        for i in range(1, n + 1):
            for j in range(1, m + 1):
                score = self.calculate_similarity(ch_syllables[i-1], in_syllables[j-1])
                
                # 計算三種可能，但 delete (垂直) 設為極低分以禁用
                match = dp[i-1][j-1] + score
                insert = dp[i][j-1] + gap_penalty
                delete = -9999.0 # 禁止垂直方向 (↑)
                
                best_score = max(match, insert, delete)
                dp[i][j] = best_score
                
                if best_score == insert:
                    dir_matrix[i][j] = "←"
                elif best_score == match:
                    dir_matrix[i][j] = "↖"
                else:
                    dir_matrix[i][j] = "↑" # 理論上不會走到這，除非所有路徑都不可行
        
        i, j = n, m
        alignment = []
        path_coords = [(i, j)] 
        
        while i > 0 or j > 0:
            current_score = dp[i][j]
            
            score_match = -9999.0
            if i > 0 and j > 0:
                sim = self.calculate_similarity(ch_syllables[i-1], in_syllables[j-1])
                score_match = dp[i-1][j-1] + sim
            
            score_ins = dp[i][j-1] + gap_penalty if j > 0 else -9999.0
            # 移除 score_del 檢查以符合「不能垂直」的要求
            
            # 優先檢查 Insert (水平)
            if j > 0 and abs(current_score - score_ins) < 1e-5:
                alignment.append((None, in_syllables[j-1], "中文缺失"))
                j -= 1
                
            # 其次檢查 Match (對角線)
            elif i > 0 and j > 0 and abs(current_score - score_match) < 1e-5:
                alignment.append((ch_syllables[i-1], in_syllables[j-1], "已匹配"))
                i -= 1; j -= 1
            
            # 如果發生死路（理論上在只能水平/斜向且要消耗完所有 i 的情況下，若 m < n 會發生）
            else:
                if i > 0: i -= 1 # 強制消耗 i 防止死迴圈，但標註缺失
                if j > 0: j -= 1
                
            path_coords.append((i, j))
        
        return alignment[::-1], dp, dir_matrix, path_coords

    def merge_syllables(self, syl_base, syl_append):
        middle_part = syl_append['initial']
        if middle_part in ['0c', '0v']: 
            middle_part = ''
        
        new_final = syl_base['final'] + middle_part + syl_append['final']
        new_pinyin = syl_base.get('pinyin', '') + syl_append.get('pinyin', '')
        return {
            'initial': syl_base['initial'],
            'final': new_final,
            'pinyin': new_pinyin
        }

    def refine_alignment(self, original_alignment):
        current_alignment = original_alignment
        merge_history = [] 
        
        while True:
            has_changed = False
            new_alignment = []
            n = len(current_alignment)
            i = 0
            while i < n:
                ch_curr, ts_curr, status = current_alignment[i]
                if status == "中文缺失" and ts_curr is not None:
                    orphan = ts_curr
                    
                    # 初始化變數
                    delta_left = -float('inf')
                    merged_left_ts = None
                    left_candidate_str = "無"
                    
                    delta_right = -float('inf')
                    merged_right_ts = None
                    right_candidate_str = "無"
                    
                    # --- 1. 嘗試與左邊合併 ---
                    if len(new_alignment) > 0 and new_alignment[-1][2] in ["已匹配", "已匹配(合併)"]:
                        prev_ch = new_alignment[-1][0]
                        prev_ts = new_alignment[-1][1]
                        left_candidate_str = prev_ts['pinyin']
                        
                        score_orig = self.calculate_similarity(prev_ch, prev_ts)
                        temp_merged = self.merge_syllables(prev_ts, orphan)
                        score_new = self.calculate_similarity(prev_ch, temp_merged)
                        
                        if(score_new >= 2 and score_new > score_orig): 
                            delta_left = score_new - score_orig + 0.5
                        else:
                            delta_left = score_new - score_orig
                        merged_left_ts = temp_merged
                        
                    # --- 2. 嘗試與右邊合併 ---
                    if i + 1 < n and current_alignment[i+1][2] in ["已匹配", "已匹配(合併)"]:
                        next_ch = current_alignment[i+1][0]
                        next_ts = current_alignment[i+1][1]
                        right_candidate_str = next_ts['pinyin']
                        
                        score_orig = self.calculate_similarity(next_ch, next_ts)
                        temp_merged = self.merge_syllables(orphan, next_ts)
                        score_new = self.calculate_similarity(next_ch, temp_merged)
                        
                        if(score_new >= 2 and score_new > score_orig): 
                            delta_right = score_new - score_orig + 0.5
                        else:
                            delta_right = score_new - score_orig
                        merged_right_ts = temp_merged
                    
                    # --- 3. 準備紀錄物件 ---
                    log_entry = {
                        "orphan_syllable": orphan['pinyin'],
                        "left_candidate": left_candidate_str,
                        "delta_left": delta_left,
                        "right_candidate": right_candidate_str,
                        "delta_right": delta_right,
                        "decision": "不合併",
                        "merged_result": "---"
                    }

                    # --- 4. 決策邏輯 ---
                    if delta_left == -float('inf') and delta_right == -float('inf'):
                        log_entry["decision"] = "無法合併(無對象)"
                        merge_history.append(log_entry) # 紀錄嘗試
                        new_alignment.append((ch_curr, ts_curr, status))
                        i += 1
                        continue
                        
                    if delta_left >= delta_right:
                        # 決定向左
                        log_entry["decision"] = "向左合併"
                        log_entry["merged_result"] = merged_left_ts['pinyin']
                        merge_history.append(log_entry)
                        
                        new_alignment[-1] = (new_alignment[-1][0], merged_left_ts, "已匹配(合併)")
                        has_changed = True
                        i += 1
                    else:
                        # 決定向右
                        log_entry["decision"] = "向右合併"
                        log_entry["merged_result"] = merged_right_ts['pinyin']
                        merge_history.append(log_entry)
                        
                        target_next = current_alignment[i+1]
                        current_alignment[i+1] = (target_next[0], merged_right_ts, "已匹配(合併)")
                        has_changed = True
                        i += 1
                else:
                    new_alignment.append((ch_curr, ts_curr, status))
                    i += 1
            current_alignment = new_alignment
            if not has_changed:
                break
        return current_alignment, merge_history

def parse_pinyin_string(pinyin_str):
    if not pinyin_str:
        return {'initial': '', 'final': '', 'pinyin': ''}
    parts = pinyin_str.split(',')
    if len(parts) == 2:
        initial, final = parts[0].strip(), parts[1].strip()
    else:
        initial, final = "", pinyin_str.strip()
    return {'initial': initial, 'final': final, 'pinyin': initial + final}

def process_single_file(json_file_path, output_excel_path, output_json_path):
    """
    處理單一檔案的函式
    """
    if not os.path.exists(json_file_path):
        print(f"❌ 找不到檔案 {json_file_path}")
        return

    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    aligner = PhoneticAligner()
    wb = Workbook()
    
    # --- Sheet 1: 對齊列表 ---
    ws_summary = wb.active
    ws_summary.title = "對齊結果列表"
    headers = ["詞彙", "語意", "標準拼音", "實際拼音", "對齊視覺化", "狀態", "詳細分數"]
    ws_summary.append(headers)
    
    # --- Sheet 2: 矩陣視覺化 ---
    ws_matrix = wb.create_sheet("矩陣視覺化")
    
    # --- Sheet 3: 合併過程紀錄 (新功能 - 詳細版) ---
    ws_merge_log = wb.create_sheet("合併過程紀錄")
    # [修改] 欄位增加左右比較
    ws_merge_log.append(["詞彙", "孤兒音節", "左側對象", "左側分數增益", "右側對象", "右側分數增益", "最終決策", "合併後結果"])
    
    header_font = Font(bold=True)
    for cell in ws_merge_log[1]:
        cell.font = header_font
    
    red_bold_font = Font(color="FF0000", bold=True)
    highlight_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    current_matrix_row = 1
    json_results = []

    def clean_syllable_data(syl_dict):
        if not syl_dict: return None
        new_dict = syl_dict.copy()
        new_dict['initial'] = new_dict['initial'].replace("0c", "").replace("0v", "")
        new_dict['final'] = new_dict['final'].replace("0c", "").replace("0v", "")
        new_dict['pinyin'] = new_dict['pinyin'].replace("0c", "").replace("0v", "")
        return new_dict

    def format_score(val):
        if val == -float('inf'): return "不可行"
        return f"{val:.4f}"

    def clean_str(s):
        return s.replace("0c", "").replace("0v", "")

    for entry in data:
        raw_dict_pinyin = entry.get('dict_pinyin', [])
        raw_pinyin_info = entry.get('pinyin_info', [])
        chinese_word = entry.get('chinese', '') 
        
        ch_input = [parse_pinyin_string(s) for s in raw_dict_pinyin]
        ts_input = [parse_pinyin_string(s) for s in raw_pinyin_info]

        alignment_result, dp_matrix, dir_matrix, path_coords = aligner.align(ch_input, ts_input)
        
        final_alignment, merge_history = aligner.refine_alignment(alignment_result)
        
        # [修改] 寫入詳細合併紀錄
        if merge_history:
            for log in merge_history:
                ws_merge_log.append([
                    chinese_word,
                    clean_str(log['orphan_syllable']),
                    clean_str(log['left_candidate']),
                    format_score(log['delta_left']),
                    clean_str(log['right_candidate']),
                    format_score(log['delta_right']),
                    log['decision'],
                    clean_str(log['merged_result'])
                ])
        
        aligned_syllables_data = []
        aligned_pairs = ["|"]
        status_list = []
        details_list = []
        
        for ch, ts, status in final_alignment:
            c_str = ch['pinyin'] if ch else "---"
            t_str = ts['pinyin'] if ts else "---"
            c_str = c_str.replace("0c", "").replace("0v", "")
            t_str = t_str.replace("0c", "").replace("0v", "")
            
            aligned_pairs.append(f"{c_str} ↔ {t_str} | ")
            status_list.append(status)
            
            if ch and ts:
                onset_score = aligner.calculate_consonant_score(ch['initial'], ts['initial'])
                dice_score = aligner.dice_coefficient(ch['final'], ts['final'])
                details_list.append(f"聲:{'Yes' if onset_score>0 else 'No'} | 韻:{dice_score:.2f}")
            else:
                details_list.append("---")
            
            ch_data = ch
            ts_data = ts
            if status in ["已匹配", "已匹配(合併)"]:
                ch_data = clean_syllable_data(ch)
                ts_data = clean_syllable_data(ts)

            syllable_entry = {
                "status": status,
                "chinese_syllable": ch_data if ch_data else None,
                "tsou_syllable": ts_data if ts_data else None
            }
            aligned_syllables_data.append(syllable_entry)

        tsou_word_reconstructed = aligner.reconstruct_tsou_word(final_alignment)
        
        json_results.append({
            "chinese": entry.get('chinese', ''),
            "ch_semantic": entry.get('ch_semantic', ''),
            "original_pinyin_ch": raw_dict_pinyin,
            "original_pinyin_ts": raw_pinyin_info,
            "original_tsou_word": tsou_word_reconstructed,
            "alignment": aligned_syllables_data
        })

        ws_summary.append([
            entry.get('chinese', ''),
            entry.get('ch_semantic', ''),
            "-".join([str(s) if s else "" for s in raw_dict_pinyin]),
            "-".join([str(s) if s else "" for s in raw_pinyin_info]),
            "\n".join(aligned_pairs),
            ", ".join(status_list),
            "\n".join(details_list)
        ])

        # 矩陣視覺化部分
        ws_matrix.cell(row=current_matrix_row, column=1, value=f"詞彙: {entry.get('chinese', '')}")
        ws_matrix.cell(row=current_matrix_row, column=1).font = Font(bold=True, size=12)
        current_matrix_row += 1

        row_headers = ["Start"] + [x['pinyin'] for x in ch_input]
        col_headers = ["Start"] + [x['pinyin'] for x in ts_input]

        start_col_idx = 2
        for idx, text in enumerate(col_headers):
            cell = ws_matrix.cell(row=current_matrix_row, column=start_col_idx + idx, value=text)
            cell.alignment = center_align
            cell.font = Font(bold=True)
            cell.border = thin_border

        current_matrix_row += 1
        n_rows, m_cols = dp_matrix.shape
        for i in range(n_rows):
            header_cell = ws_matrix.cell(row=current_matrix_row + i, column=1, value=row_headers[i])
            header_cell.alignment = center_align
            header_cell.font = Font(bold=True)
            header_cell.border = thin_border

            for j in range(m_cols):
                val = dp_matrix[i][j]
                direction = dir_matrix[i][j]
                display_text = f"{direction} {val:.2f}" if direction else f"{val:.2f}"
                cell = ws_matrix.cell(row=current_matrix_row + i, column=start_col_idx + j, value=display_text)
                cell.alignment = center_align
                cell.border = thin_border
                if (i, j) in path_coords:
                    cell.font = red_bold_font
                    cell.fill = highlight_fill
        current_matrix_row += n_rows + 2

    # 自動調整欄寬
    for ws in [ws_summary, ws_merge_log]:
        for col in ws.columns:
            lengths = [len(str(cell.value)) for cell in col if cell.value is not None]
            length = max(lengths) if lengths else 0
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 5, 50)

    try:
        wb.save(output_excel_path)
        print(f"   💾 Excel 已儲存: {os.path.basename(output_excel_path)}")
    except Exception as e:
        print(f"   ❌ Excel 儲存失敗: {e}")

    try:
        with open(output_json_path, 'w', encoding='utf-8') as f:
            json.dump(json_results, f, ensure_ascii=False, indent=4)
        print(f"   💾 JSON 已儲存: {os.path.basename(output_json_path)}")
    except Exception as e:
        print(f"   ❌ JSON 儲存失敗: {e}")

def main():
    # 檢查輸入資料夾
    if not os.path.exists(INPUT_FOLDER):
        print(f"❌ 找不到輸入資料夾: {INPUT_FOLDER}")
        print("請先執行上一步驟 (Align_syllables03.py)")
        return

    # 建立輸出資料夾
    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)
        print(f"📂 已建立輸出資料夾: {OUTPUT_FOLDER}")

    # 取得所有 JSON 檔案
    files = [f for f in os.listdir(INPUT_FOLDER) if f.endswith(".json")]
    print(f"🔍 發現 {len(files)} 個檔案，準備開始處理...\n")

    for filename in files:
        input_path = os.path.join(INPUT_FOLDER, filename)
        
        # 產生輸出檔名 (例如 aligned_Amis.json -> refined_Amis.xlsx)
        base_name = os.path.splitext(filename)[0]
        # 如果檔名有 aligned_ 前綴，可以考慮替換成 refined_，或直接加後綴
        new_base_name = base_name.replace("aligned_", "refined_") if "aligned_" in base_name else f"refined_{base_name}"
        
        output_excel = os.path.join(OUTPUT_FOLDER, f"{new_base_name}.xlsx")
        output_json = os.path.join(OUTPUT_FOLDER, f"{new_base_name}.json")

        print(f"🚀 正在處理: {filename} ...")
        process_single_file(input_path, output_excel, output_json)
        print("-" * 30)

    print("\n🎉 所有檔案處理完成！")

if __name__ == "__main__":
    main()
