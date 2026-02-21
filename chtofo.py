import pandas as pd
import os
import re
from collections import defaultdict, Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def main():
    # --- 1. 路徑設定 ---
    base_path = os.path.dirname(os.path.abspath(__file__))
    
    # 輸入檔案
    INPUT_FILE_PATH = os.path.join(base_path, 'output_alignment_refined.xlsx')
    
    # 輸出檔案 1: 總表 (每個中文只列出第一名的族語發音)
    OUTPUT_WINNER_PATH = os.path.join(base_path, 'most_frequent_summary.xlsx')
    
    # 輸出檔案 2: 詳細統計版 (含顏色標記)
    OUTPUT_STATS_COLORED_PATH = os.path.join(base_path, 'pronunciation_stats_colored.xlsx')

    print(f"正在讀取檔案: {INPUT_FILE_PATH}")

    if not os.path.exists(INPUT_FILE_PATH):
        print(f"錯誤: 找不到檔案 {INPUT_FILE_PATH}")
        return

    # --- 2. 讀取與解析 ---
    try:
        df = pd.read_excel(INPUT_FILE_PATH, sheet_name='對齊結果列表')
    except Exception as e:
        print(f"讀取 Excel 失敗: {e}")
        return

   
    stats_map = defaultdict(Counter)

    for index, row in df.iterrows():
        status = str(row['狀態'])
        visual_content = str(row['對齊視覺化'])

        # 過濾缺失或空值
        if "族語缺失" in status or visual_content == 'nan' or not visual_content.strip():
            continue

        # 解析內容
        segments = re.split(r'[|\n]', visual_content)
        for segment in segments:
            if '↔' in segment:
                parts = segment.split('↔')
                if len(parts) == 2:
                    chi = parts[0].strip()
                    ind = parts[1].strip()
                    if chi and ind and chi != '---':
                        stats_map[chi][ind] += 1

    # --- 3. 找出每個中文發音族與頻率最高 (用於檔案 1 和 檔案 2 的標記) ---
    winner_map = {} # 記錄冠軍是誰: { 'bao': 'paw' }
    winner_data = [] # 用於產生檔案 1

    for chi, counter in stats_map.items():
        # 取得第一名
        most_common = counter.most_common(1)[0]
        winner_ind = most_common[0]
        winner_count = most_common[1]
        total_count = sum(counter.values())
        
        winner_map[chi] = winner_ind
        
        winner_data.append({
            "中文發音": chi,
            "最常出現族語發音": winner_ind,
            "該發音次數": winner_count,
            "總樣本數": total_count,
            "佔比": f"{(winner_count / total_count * 100):.1f}%"
        })

    # --- 4. 產生檔案 1: 總表(出現最多) ---
    df_winner = pd.DataFrame(winner_data)
    df_winner = df_winner.sort_values(by='中文發音')
    df_winner.to_excel(OUTPUT_WINNER_PATH, index=False)
    print(f"檔案 1 (總表) 已產生: {OUTPUT_WINNER_PATH}")

    # --- 5. 產生檔案 2: 詳細統計版  ---
    detailed_data = []
    
    for chi, counter in stats_map.items():
        # 將該中文的所有對應都列出來
        for ind, count in counter.items():
            is_winner = (ind == winner_map[chi])
            detailed_data.append({
                "中文發音": chi,
                "族語發音": ind,
                "出現次數": count,
                "是否為最常出現": "是" if is_winner else "否" # 輔助欄位，方便程式判斷顏色
            })

    df_detail = pd.DataFrame(detailed_data)
    # 排序: 中文 A-Z -> 次數 多-少
    df_detail = df_detail.sort_values(by=['中文發音', '出現次數'], ascending=[True, False])
    
    # 先存成 Excel，再用 openpyxl 上色
    df_detail.to_excel(OUTPUT_STATS_COLORED_PATH, index=False)

    # --- 6. 檔案 2 後製: 上色 ---
    print("正在為統計版上色...")
    wb = load_workbook(OUTPUT_STATS_COLORED_PATH)
    ws = wb.active # 預設第一個 sheet

    # 定義顏色: 淺綠色 (標記)
    fill_winner = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # 取得欄位索引
    header = {cell.value: i for i, cell in enumerate(ws[1])}
    col_is_winner = header.get('是否為最常出現')

    if col_is_winner is not None:
        # 遍歷每一列 (跳過標題)
        for row in ws.iter_rows(min_row=2):
            # 檢查 "是否為最常出現" 欄位
            cell_check = row[col_is_winner]
            
            if cell_check.value == "是":
                # 如果是最多，整列上色
                for cell in row:
                    cell.fill = fill_winner
    
    wb.save(OUTPUT_STATS_COLORED_PATH)
    print(f"檔案 2 (詳細統計版+顏色) 已產生: {OUTPUT_STATS_COLORED_PATH}")
    print("完成！綠色底色代表該發音是該中文對應中最常出現的。")

if __name__ == "__main__":
    main()
